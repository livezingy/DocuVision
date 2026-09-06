"""Unit tests for ExportService table banners (CSV / Markdown / XLSX).

No Paddle / GPU. CSV and Markdown are pure logic; XLSX is skipped when
pandas/openpyxl are unavailable.
"""

from __future__ import annotations

import asyncio
from typing import Any, Dict, List

import pytest

from app.services.export_service import (
    ExportService,
    excel_safe_cell,
    format_table_csv_banner,
    format_table_export_title,
    table_confidence_pct,
)


def _table(
    *,
    page: Any = 16,
    confidence: Any = None,
    score: Any = None,
    data: List[List[str]] | None = None,
    rows: int = 2,
    columns: int = 2,
) -> Dict[str, Any]:
    item: Dict[str, Any] = {
        "page": page,
        "data": data if data is not None else [["H1", "H2"], ["a", "b"]],
        "rows": rows,
        "columns": columns,
    }
    if confidence is not None:
        item["confidence"] = confidence
    if score is not None:
        item["score"] = score
    return item


class TestTableConfidencePct:
    def test_ratio_rounded_to_percent(self):
        assert table_confidence_pct({"confidence": 0.87}) == 87

    def test_one_is_one_hundred(self):
        assert table_confidence_pct({"confidence": 1.0}) == 100

    def test_zero_is_zero_not_fallback(self):
        assert table_confidence_pct({"confidence": 0.0, "score": 0.9}) == 0

    def test_missing_confidence_falls_back_to_score(self):
        assert table_confidence_pct({"score": 0.5}) == 50

    def test_already_percent_score(self):
        assert table_confidence_pct({"score": 92}) == 92

    def test_neither_is_zero(self):
        assert table_confidence_pct({}) == 0

    def test_invalid_is_zero(self):
        assert table_confidence_pct({"confidence": "n/a"}) == 0


class TestTableBannerFormat:
    def test_title_and_csv_banner(self):
        table = _table(page=16, confidence=0.87)
        assert format_table_export_title(7, table) == "Table 7 (Page 16) confidence=87%"
        assert format_table_csv_banner(7, table) == "=== Table 7 (Page 16) confidence=87% ==="

    def test_score_fallback_in_banner(self):
        table = _table(page=3, score=0.5)
        assert format_table_csv_banner(2, table) == "=== Table 2 (Page 3) confidence=50% ==="

    def test_title_appends_caption(self):
        table = _table(page=16, confidence=0.87)
        table["caption"] = "Table 1: Ablation results"
        assert format_table_export_title(1, table) == (
            "Table 1 (Page 16) confidence=87% — Table 1: Ablation results"
        )
        assert format_table_csv_banner(1, table) == "=== Table 1 (Page 16) confidence=87% ==="


class TestExcelSafeCell:
    def test_plus_prefixed_text(self):
        assert excel_safe_cell("+Accuracy") == "'+Accuracy"

    def test_equals_formula(self):
        assert excel_safe_cell("=SUM(1)") == "'=SUM(1)"

    def test_numeric_minus_kept(self):
        assert excel_safe_cell("-0.5") == "-0.5"

    def test_text_minus_prefixed(self):
        assert excel_safe_cell("-Accuracy") == "'-Accuracy"

    def test_plain_unchanged(self):
        assert excel_safe_cell("hello") == "hello"


class TestExportCsvAndMarkdown:
    def test_csv_banner_per_table(self, tmp_path, monkeypatch):
        from app.core.config import settings as _settings

        monkeypatch.setattr(_settings, "OUTPUT_DIR", str(tmp_path))
        svc = ExportService()
        result = {
            "tables": [
                _table(page=16, confidence=0.87),
                _table(page=3, score=0.5, data=[["X"]], rows=1, columns=1),
            ]
        }
        path = asyncio.run(svc.to_csv(result, "t1"))
        text = open(path, encoding="utf-8-sig").read()
        assert "=== Table 1 (Page 16) confidence=87% ===" in text
        assert "=== Table 2 (Page 3) confidence=50% ===" in text
        assert "H1" in text and "a" in text

    def test_csv_escapes_plus_and_writes_caption(self, tmp_path, monkeypatch):
        from app.core.config import settings as _settings

        monkeypatch.setattr(_settings, "OUTPUT_DIR", str(tmp_path))
        svc = ExportService()
        table = _table(page=16, confidence=0.87, data=[["H", "+pos"], ["a", "b"]])
        table["caption"] = "Table 1: Ablation results"
        path = asyncio.run(svc.to_csv({"tables": [table]}, "t-plus"))
        text = open(path, encoding="utf-8-sig").read()
        assert "=== Table 1 (Page 16) confidence=87% ===" in text
        assert "Caption: Table 1: Ablation results" in text
        assert "'+pos" in text
        assert "### " not in text

    def test_markdown_heading_includes_confidence(self, tmp_path, monkeypatch):
        from app.core.config import settings as _settings

        monkeypatch.setattr(_settings, "OUTPUT_DIR", str(tmp_path))
        svc = ExportService()
        md = asyncio.run(svc.to_markdown({"tables": [_table(page=16, confidence=0.87)]}))
        assert "### Table 1 (Page 16) confidence=87%" in md

    def test_markdown_heading_includes_caption(self, tmp_path, monkeypatch):
        from app.core.config import settings as _settings

        monkeypatch.setattr(_settings, "OUTPUT_DIR", str(tmp_path))
        svc = ExportService()
        table = _table(page=16, confidence=0.87)
        table["caption"] = "Table 1: Ablation results"
        md = asyncio.run(svc.to_markdown({"tables": [table]}))
        assert "### Table 1 (Page 16) confidence=87% — Table 1: Ablation results" in md


class TestExportExcel:
    def test_xlsx_title_row(self, tmp_path, monkeypatch):
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        from app.core.config import settings as _settings

        monkeypatch.setattr(_settings, "OUTPUT_DIR", str(tmp_path))
        svc = ExportService()
        result = {"document_info": {}, "tables": [_table(page=16, confidence=0.87)]}
        path = asyncio.run(svc.to_excel(result, "t-xlsx"))
        wb = load_workbook(path)
        ws = wb["Table_1"]
        assert ws.cell(row=1, column=1).value == "Table 1 (Page 16) confidence=87%"
        assert ws.cell(row=2, column=1).value == "H1"
        assert ws.cell(row=3, column=1).value == "a"
